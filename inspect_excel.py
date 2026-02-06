import openpyxl

wb = openpyxl.load_workbook('Book1.xlsx')
ws = wb.active

# Print first row (headers)
print("Headers:")
headers = []
for cell in ws[1]:
    print(f"Column {cell.column}: {cell.value}")
    headers.append(cell.value)

print("\n\nFirst product data row:")
row_data = list(ws[2])
for i, cell in enumerate(row_data):
    print(f"Column {i}: {cell.value}")
